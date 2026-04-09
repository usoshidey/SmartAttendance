"""
tasks.py
Celery background tasks.
Speed improvements vs original:
  - Face detector reads video directly (no frame-to-disk step)
  - Frame resizing before YOLO
  - sample_rate=5 (skip 4 of 5 frames)
Everything else — filtering, embeddings, DBSCAN, matching — identical to original.
"""
import os
import json
import shutil
from datetime import datetime, date
from pathlib import Path
from celery import Celery

from backend.config import (
    REDIS_URL, UPLOADS_DIR, ATTENDANCE_REPORTS_DIR,
    YOLO_MODEL_PATH, FRAME_SAMPLE_RATE,
    MIN_FACE_SIZE, BLUR_THRESHOLD, YOLO_CONF, DEEPFACE_MODEL,
    DBSCAN_EPS_VALUES,
    SIMILARITY_THRESHOLD, DATABASE_URL
)
from backend.pipeline.face_detector import detect_and_crop_faces
from backend.pipeline.face_filter import filter_faces
from backend.pipeline.embedder import generate_embeddings
from backend.pipeline.clusterer import cluster_faces
from backend.pipeline.matcher import match_clusters_to_students

celery_app = Celery("smart_attendance", broker=REDIS_URL, backend=REDIS_URL)
celery_app.conf.task_serializer = "json"
celery_app.conf.result_serializer = "json"
celery_app.conf.accept_content = ["json"]
celery_app.conf.broker_connection_retry_on_startup = True


def _get_db():
    from sqlalchemy import create_engine
    from sqlalchemy.orm import sessionmaker
    engine = create_engine(
        DATABASE_URL,
        connect_args={"check_same_thread": False} if "sqlite" in DATABASE_URL else {}
    )
    return sessionmaker(bind=engine)()


def _update_job(db, job_id, status=None, progress=None, message=None,
                result_path=None, error=None):
    from backend.database import Job
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        return
    if status:
        job.status = status
        if status in ("done", "failed"):
            job.completed_at = datetime.utcnow()
    if progress is not None:
        job.progress = progress
    if message:
        job.progress_message = message
    if result_path:
        job.result_path = result_path
    if error:
        job.error_message = error
    db.commit()


# ── REGISTRATION TASK ──────────────────────────────────────────────────────────

@celery_app.task(bind=True)
def run_registration_pipeline(self, job_id: str):
    db = _get_db()

    def update(pct, msg):
        _update_job(db, job_id, progress=pct, message=msg)

    try:
        from backend.database import Job, ClusterResult
        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        _update_job(db, job_id, status="processing", progress=0,
                    message="Starting pipeline...")

        # Rename folder from UUID → meaningful name: subjectcode_date_shortid
        from backend.database import Subject as _Subject
        _subj      = db.query(_Subject).filter(_Subject.id == job.subject_id).first()
        _subj_code = (_subj.code if _subj else "unknown").replace(" ", "_")
        _subj_name = (_subj.name if _subj else "unknown").replace(" ", "_")
        _date_str  = date.today().strftime("%Y-%m-%d")
        folder_name = f"{_subj_code}_{_date_str}_{job_id[:8]}"
        old_job_dir = str(UPLOADS_DIR / job_id)
        job_dir     = str(UPLOADS_DIR / folder_name)

        # Rename the UUID folder that jobs.py created
        if os.path.exists(old_job_dir) and not os.path.exists(job_dir):
            os.rename(old_job_dir, job_dir)
        elif not os.path.exists(job_dir):
            os.makedirs(job_dir, exist_ok=True)

        # Update video_path in DB so downstream steps find the video
        new_video_path = job.video_path.replace(old_job_dir, job_dir)
        job.video_path = new_video_path
        db.commit()

        faces_dir      = os.path.join(job_dir, "faces")
        filtered_dir   = os.path.join(job_dir, "faces_filtered")
        emb_path       = os.path.join(job_dir, "embeddings.json")
        os.makedirs(faces_dir,    exist_ok=True)
        os.makedirs(filtered_dir, exist_ok=True)

        # ── Step 1: Video → Face crops (direct, no frame saving) ─────────────
        def p1(pct, msg):
            update(int(pct * 0.30), f"[1/4] {msg}")

        update(1, "[1/4] Loading video and running YOLO face detection...")
        detect_and_crop_faces(
            frames_dir=None,
            output_dir=faces_dir,
            model_path=str(YOLO_MODEL_PATH),
            conf=YOLO_CONF,
            progress_callback=p1,
            video_path=job.video_path,
            sample_rate=FRAME_SAMPLE_RATE,
            resize_width=960,
        )

        # ── Step 2: Filter faces (multi-stage quality check) ─────────────────
        def p2(pct, msg):
            update(30 + int(pct * 0.10), f"[2/4] {msg}")
        filter_faces(
            faces_dir, filtered_dir,
            min_face_size=MIN_FACE_SIZE,
            blur_threshold=BLUR_THRESHOLD,
            progress_callback=p2
        )

        # ── Step 3: Generate embeddings (VGG-Face) ────────────────────────────
        def p3(pct, msg):
            update(40 + int(pct * 0.35), f"[3/4] {msg}")
        embeddings = generate_embeddings(filtered_dir, emb_path, DEEPFACE_MODEL, p3)

        if not embeddings:
            raise ValueError(
                "No embeddings generated. Check video quality or YOLO model path."
            )

        # ── Step 4: DBSCAN — dynamic per-class grid search ───────────────────
        def p4(pct, msg):
            update(75 + int(pct * 0.20), f"[4/4] {msg}")
        clusters = cluster_faces(
            embeddings,
            eps_values=DBSCAN_EPS_VALUES,
            min_samples_values=None,   # ← None = computed dynamically per class
            progress_callback=p4
        )

        # ── Save cluster results to DB ────────────────────────────────────────
        db.query(ClusterResult).filter(ClusterResult.job_id == job_id).delete()
        clusters_vis_dir = os.path.join(job_dir, "clusters_vis")

        for cid, data in clusters.items():
            cluster_img_dir = os.path.join(clusters_vis_dir, f"cluster_{cid}")
            os.makedirs(cluster_img_dir, exist_ok=True)

            # Score all faces by sharpness + brightness, pick top 6 best quality
            import cv2 as _cv2
            import numpy as _np
            scored = []
            for fname in data["files"]:
                src = os.path.join(filtered_dir, fname)
                if not os.path.exists(src):
                    continue
                img = _cv2.imread(src, _cv2.IMREAD_GRAYSCALE)
                if img is None:
                    continue
                sharpness  = _cv2.Laplacian(img, _cv2.CV_64F).var()
                brightness = float(_np.mean(img))
                # Penalise too dark (<50) or too bright (>210)
                bright_ok  = 1.0 if 50 <= brightness <= 210 else 0.3
                score      = sharpness * bright_ok
                scored.append((score, fname, src))

            # Sort descending by quality, take best 6
            scored.sort(key=lambda x: x[0], reverse=True)
            best_faces = scored[:6]

            face_web_paths = []
            for _, fname, src in best_faces:
                dst = os.path.join(cluster_img_dir, fname)
                shutil.copy(src, dst)
                face_web_paths.append(dst)

            db.add(ClusterResult(
                job_id=job_id,
                cluster_id=int(cid),
                embedding=json.dumps(data["centroid"]),
                face_paths=json.dumps(face_web_paths)
            ))

        db.commit()
        _update_job(db, job_id, status="done", progress=100,
                    message=f"Done! {len(clusters)} student clusters found. "
                            f"Please assign names.")

    except Exception as e:
        _update_job(db, job_id, status="failed", progress=0,
                    message="Pipeline failed.", error=str(e))
        raise
    finally:
        db.close()


# ── ATTENDANCE TASK ────────────────────────────────────────────────────────────

@celery_app.task(bind=True)
def run_attendance_pipeline(self, job_id: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    db = _get_db()

    def update(pct, msg):
        _update_job(db, job_id, progress=pct, message=msg)

    try:
        from backend.database import Job, Student, AttendanceSession, AttendanceRecord

        job = db.query(Job).filter(Job.id == job_id).first()
        if not job:
            return

        _update_job(db, job_id, status="processing", progress=0,
                    message="Starting pipeline...")

        from backend.database import Subject as _Subject
        _subj      = db.query(_Subject).filter(_Subject.id == job.subject_id).first()
        _subj_code = (_subj.code if _subj else "unknown").replace(" ", "_")
        _subj_name = (_subj.name if _subj else "unknown").replace(" ", "_")
        _date_str  = date.today().strftime("%Y-%m-%d")
        folder_name = f"{_subj_code}_{_date_str}_{job_id[:8]}"
        old_job_dir = str(UPLOADS_DIR / job_id)
        job_dir     = str(UPLOADS_DIR / folder_name)

        if os.path.exists(old_job_dir) and not os.path.exists(job_dir):
            os.rename(old_job_dir, job_dir)
        elif not os.path.exists(job_dir):
            os.makedirs(job_dir, exist_ok=True)

        new_video_path = job.video_path.replace(old_job_dir, job_dir)
        job.video_path = new_video_path
        db.commit()

        faces_dir    = os.path.join(job_dir, "faces")
        filtered_dir = os.path.join(job_dir, "faces_filtered")
        emb_path     = os.path.join(job_dir, "embeddings.json")
        os.makedirs(faces_dir,    exist_ok=True)
        os.makedirs(filtered_dir, exist_ok=True)

        # ── Step 1: Video → Face crops ────────────────────────────────────────
        def p1(pct, msg):
            update(int(pct * 0.30), f"[1/4] {msg}")

        update(1, "[1/4] Loading video and running YOLO face detection...")
        detect_and_crop_faces(
            frames_dir=None,
            output_dir=faces_dir,
            model_path=str(YOLO_MODEL_PATH),
            conf=YOLO_CONF,
            progress_callback=p1,
            video_path=job.video_path,
            sample_rate=FRAME_SAMPLE_RATE,
            resize_width=960,
        )

        # ── Step 2: Filter ────────────────────────────────────────────────────
        def p2(pct, msg):
            update(30 + int(pct * 0.10), f"[2/4] {msg}")
        filter_faces(
            faces_dir, filtered_dir,
            min_face_size=MIN_FACE_SIZE,
            blur_threshold=BLUR_THRESHOLD,
            progress_callback=p2
        )

        # ── Step 3: Embeddings ────────────────────────────────────────────────
        def p3(pct, msg):
            update(40 + int(pct * 0.35), f"[3/4] {msg}")
        embeddings = generate_embeddings(filtered_dir, emb_path, DEEPFACE_MODEL, p3)

        # ── Step 4: Cluster — dynamic per-class grid search ───────────────────
        def p4(pct, msg):
            update(75 + int(pct * 0.10), f"[4/4] Clustering... {msg}")
        clusters = cluster_faces(
            embeddings,
            eps_values=DBSCAN_EPS_VALUES,
            min_samples_values=None,   # ← computed dynamically per class
            progress_callback=p4
        )

        # ── Step 5: Match against registered students ─────────────────────────
        update(85, "Matching faces to registered students...")
        registered = db.query(Student).filter(
            Student.subject_id == job.subject_id
        ).all()
        registered_data = [
            {"student_id": s.id, "name": s.name, "roll_no": s.roll_no,
             "embedding": s.get_embedding()}
            for s in registered
        ]
        matches = match_clusters_to_students(
            clusters, registered_data, SIMILARITY_THRESHOLD
        )

        # ── Step 6: Save attendance to DB ─────────────────────────────────────
        update(90, "Saving attendance records...")
        session = AttendanceSession(
            job_id=job_id, subject_id=job.subject_id, date=date.today()
        )
        db.add(session)
        db.flush()

        for m in matches:
            db.add(AttendanceRecord(
                session_id=session.id,
                student_id=m["student_id"],
                status=m["status"],
                similarity=m["similarity"]
            ))
        db.commit()

        # ── Step 7: Apply 50% similarity hard threshold ──────────────────────
        # Any "present" student with similarity < 0.50 → forced absent
        HARD_THRESHOLD = 0.50
        for m in matches:
            if m["status"] == "present" and (
                m["similarity"] is None or m["similarity"] < HARD_THRESHOLD
            ):
                m["status"] = "absent"

        # Re-save attendance records with corrected statuses
        for m in matches:
            record = db.query(AttendanceRecord).filter(
                AttendanceRecord.session_id == session.id,
                AttendanceRecord.student_id == m["student_id"]
            ).first()
            if record:
                record.status = m["status"]
        db.commit()

        # ── Step 8: Excel report with face images ─────────────────────────────
        from openpyxl.drawing.image import Image as XLImage
        from PIL import Image as PILImage
        import io

        update(95, "Generating Excel report with face images...")
        report_path = str(ATTENDANCE_REPORTS_DIR / f"{_subj_name}_{_date_str}.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance"

        # Build a lookup: student_id → sample face path
        student_face_map = {}
        for s in registered:
            if s.sample_face_path and os.path.exists(s.sample_face_path):
                student_face_map[s.id] = s.sample_face_path

        headers = ["Roll No.", "Name", "Face", "Status", "Similarity %"]
        hfill = PatternFill("solid", fgColor="1a1a2e")
        hfont = Font(color="FFFFFF", bold=True, size=12)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = hfill
            cell.font = hfont
            cell.alignment = Alignment(horizontal="center")

        # Column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 14

        matches.sort(key=lambda x: (0 if x["status"] == "present" else 1, x["roll_no"]))
        pfill = PatternFill("solid", fgColor="d4edda")
        afill = PatternFill("solid", fgColor="f8d7da")
        ROW_HEIGHT = 55

        for row, m in enumerate(matches, 2):
            sim_pct = m["similarity"] * 100 if m["similarity"] else 0
            sim_str = f"{sim_pct:.1f}%" if m["similarity"] else "—"
            fill    = pfill if m["status"] == "present" else afill

            ws.cell(row=row, column=1, value=m["roll_no"]).fill = fill
            ws.cell(row=row, column=2, value=m["name"]).fill    = fill
            ws.cell(row=row, column=4, value=m["status"].title()).fill = fill
            ws.cell(row=row, column=5, value=sim_str).fill      = fill
            for col in [1, 2, 4, 5]:
                ws.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            # Insert face image in column C
            face_path = student_face_map.get(m["student_id"])
            if face_path:
                try:
                    pil_img = PILImage.open(face_path).convert("RGB")
                    pil_img.thumbnail((60, 60))
                    img_bytes = io.BytesIO()
                    pil_img.save(img_bytes, format="PNG")
                    img_bytes.seek(0)
                    xl_img = XLImage(img_bytes)
                    xl_img.width  = 50
                    xl_img.height = 50
                    ws.add_image(xl_img, f"C{row}")
                except Exception:
                    ws.cell(row=row, column=3, value="—")

            ws.row_dimensions[row].height = ROW_HEIGHT

        wb.save(report_path)

        present_count = sum(1 for m in matches if m["status"] == "present")
        _update_job(db, job_id, status="done", progress=100,
                    message=f"Done! Present: {present_count}/{len(matches)}",
                    result_path=report_path)

    except Exception as e:
        _update_job(db, job_id, status="failed", progress=0,
                    message="Pipeline failed.", error=str(e))
        raise
    finally:
        db.close()
