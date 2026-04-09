"""
routers/attendance.py
Fetch attendance session results and download Excel reports.
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session

from backend.database import get_db, Job, AttendanceSession, AttendanceRecord, Student
from backend.models.schemas import AttendanceSessionOut, AttendanceRecordOut

router = APIRouter(prefix="/attendance", tags=["Attendance"])


@router.get("/session/{job_id}", response_model=AttendanceSessionOut)
def get_attendance_by_job(job_id: str, db: Session = Depends(get_db)):
    """Get attendance results for a completed attendance job."""
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    if job.status != "done":
        raise HTTPException(400, f"Job not complete (status: {job.status})")
    if job.type != "attendance":
        raise HTTPException(400, "Not an attendance job")

    session = db.query(AttendanceSession).filter(
        AttendanceSession.job_id == job_id
    ).first()
    if not session:
        raise HTTPException(404, "Attendance session not found")

    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.session_id == session.id
    ).all()

    record_outs = []
    for r in records:
        student = db.query(Student).filter(Student.id == r.student_id).first()
        record_outs.append(AttendanceRecordOut(
            student_id=r.student_id,
            name=student.name if student else "Unknown",
            roll_no=student.roll_no if student else "—",
            status=r.status,
            similarity=r.similarity
        ))

    # Sort: present first, then by roll_no
    record_outs.sort(key=lambda x: (0 if x.status == "present" else 1, x.roll_no))

    return AttendanceSessionOut(
        session_id=session.id,
        subject_id=session.subject_id,
        date=session.date,
        records=record_outs
    )


@router.get("/download/{job_id}")
def download_attendance_report(job_id: str, db: Session = Depends(get_db)):
    """Download the Excel attendance report for a job."""
    job = db.query(Job).filter(Job.id == job_id).first()
    if not job:
        raise HTTPException(404, "Job not found")
    if not job.result_path or job.status != "done":
        raise HTTPException(400, "Report not available yet")

    import os
    if not os.path.exists(job.result_path):
        raise HTTPException(404, "Report file missing from disk")

    return FileResponse(
        job.result_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"attendance_{job_id[:8]}.xlsx"
    )


@router.get("/history/{subject_id}", response_model=list[AttendanceSessionOut])
def get_attendance_history(
    subject_id: int,
    limit: int = 10,
    db: Session = Depends(get_db)
):
    """Get past attendance sessions for a subject."""
    sessions = db.query(AttendanceSession).filter(
        AttendanceSession.subject_id == subject_id
    ).order_by(AttendanceSession.created_at.desc()).limit(limit).all()

    result = []
    for session in sessions:
        records = db.query(AttendanceRecord).filter(
            AttendanceRecord.session_id == session.id
        ).all()

        record_outs = []
        for r in records:
            student = db.query(Student).filter(Student.id == r.student_id).first()
            record_outs.append(AttendanceRecordOut(
                student_id=r.student_id,
                name=student.name if student else "Unknown",
                roll_no=student.roll_no if student else "—",
                status=r.status,
                similarity=r.similarity
            ))

        result.append(AttendanceSessionOut(
            session_id=session.id,
            subject_id=session.subject_id,
            date=session.date,
            records=record_outs
        ))
    return result


@router.get("/consolidated/{subject_id}")
def download_consolidated_report(subject_id: int, db: Session = Depends(get_db)):
    """
    Generate and download a consolidated Excel showing ALL attendance sessions
    for a subject in one sheet, plus a summary sheet.

    Sheet 1 — "Summary":
      Columns: Roll No | Name | Face | Total Classes | Present | Absent | %
      One row per student, sorted by roll_no.

    Sheet 2 — "All Sessions":
      Columns: Roll No | Name | Date1 | Date2 | Date3 | ... (P/A per date)
      One row per student, one column per session date.
      Footer row shows present count per date.
    """
    import os, io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from PIL import Image as PILImage
    from backend.database import Subject
    from backend.config import ATTENDANCE_REPORTS_DIR
    from datetime import date as date_type

    subj = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subj:
        raise HTTPException(404, "Subject not found")

    # ── Fetch all sessions for this subject sorted by date ────────────────────
    sessions = db.query(AttendanceSession).filter(
        AttendanceSession.subject_id == subject_id
    ).order_by(AttendanceSession.date.asc()).all()

    if not sessions:
        raise HTTPException(404, "No attendance sessions found for this subject")

    # ── Collect all students for this subject ─────────────────────────────────
    students = db.query(Student).filter(
        Student.subject_id == subject_id
    ).order_by(Student.roll_no).all()

    # Build: { student_id → { session_id → "P"/"A" } }
    attendance_map = {}
    for s in students:
        attendance_map[s.id] = {}

    for session in sessions:
        records = db.query(AttendanceRecord).filter(
            AttendanceRecord.session_id == session.id
        ).all()
        for r in records:
            if r.student_id in attendance_map:
                attendance_map[r.student_id][session.id] = (
                    "P" if r.status == "present" else "A"
                )

    # ── Build student face map ─────────────────────────────────────────────────
    face_map = {}
    for s in students:
        if s.sample_face_path and os.path.exists(s.sample_face_path):
            face_map[s.id] = s.sample_face_path

    # ── Excel styles ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    header_fill  = PatternFill("solid", fgColor="1a1a2e")
    header_font  = Font(color="FFFFFF", bold=True, size=11)
    present_fill = PatternFill("solid", fgColor="d4edda")
    absent_fill  = PatternFill("solid", fgColor="f8d7da")
    date_fill    = PatternFill("solid", fgColor="e8eaf6")
    summary_fill = PatternFill("solid", fgColor="fff3cd")
    center       = Alignment(horizontal="center", vertical="center")
    thin         = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC")
    )

    def styled(cell, fill=None, font=None, align=center, border=thin):
        if fill:   cell.fill      = fill
        if font:   cell.font      = font
        if align:  cell.alignment = align
        if border: cell.border    = border

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Summary"

    # Title row
    ws1.merge_cells("A1:G1")
    title_cell = ws1["A1"]
    title_cell.value = f"Attendance Summary — {subj.name} ({subj.code})"
    title_cell.font  = Font(bold=True, size=14, color="1a1a2e")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 30

    # Header row
    headers = ["Roll No.", "Name", "Face", "Total Classes", "Present", "Absent", "Attendance %"]
    col_widths = [14, 22, 10, 14, 10, 10, 14]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws1.cell(row=2, column=col, value=h)
        styled(cell, fill=header_fill, font=header_font)
        ws1.column_dimensions[cell.column_letter].width = w

    total_sessions = len(sessions)

    for row_idx, s in enumerate(students, 3):
        rec = attendance_map.get(s.id, {})
        present_count = sum(1 for v in rec.values() if v == "P")
        absent_count  = total_sessions - present_count
        pct           = round(present_count / total_sessions * 100, 1) if total_sessions > 0 else 0
        pct_fill      = PatternFill("solid", fgColor="d4edda" if pct >= 75 else ("fff3cd" if pct >= 50 else "f8d7da"))

        for col, val in enumerate([s.roll_no, s.name, "", total_sessions, present_count, absent_count, f"{pct}%"], 1):
            cell = ws1.cell(row=row_idx, column=col, value=val)
            fill = pct_fill if col >= 4 else None
            styled(cell, fill=fill)

        # Face image in column C
        fp = face_map.get(s.id)
        if fp:
            try:
                pil = PILImage.open(fp).convert("RGB")
                pil.thumbnail((50, 50))
                buf = io.BytesIO()
                pil.save(buf, format="PNG")
                buf.seek(0)
                xl = XLImage(buf)
                xl.width = 42; xl.height = 42
                ws1.add_image(xl, f"C{row_idx}")
            except Exception:
                pass

        ws1.row_dimensions[row_idx].height = 50

    # ══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — All Sessions
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("All Sessions")

    # Title
    n_date_cols = len(sessions)
    last_col_letter = openpyxl.utils.get_column_letter(3 + n_date_cols)
    ws2.merge_cells(f"A1:{last_col_letter}1")
    t2 = ws2["A1"]
    t2.value     = f"Session-wise Attendance — {subj.name} ({subj.code})"
    t2.font      = Font(bold=True, size=14, color="1a1a2e")
    t2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 30

    # Header: Roll No | Name | Date1 | Date2 | ...
    ws2.cell(row=2, column=1, value="Roll No.")
    ws2.cell(row=2, column=2, value="Name")
    for col_idx, session in enumerate(sessions, 3):
        cell = ws2.cell(row=2, column=col_idx,
                        value=session.date.strftime("%d %b %Y") if hasattr(session.date, "strftime") else str(session.date))
        styled(cell, fill=date_fill, font=Font(bold=True, size=10))

    for col in [1, 2]:
        styled(ws2.cell(row=2, column=col), fill=header_fill, font=header_font)

    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 22

    # Data rows
    for row_idx, s in enumerate(students, 3):
        ws2.cell(row=row_idx, column=1, value=s.roll_no)
        ws2.cell(row=row_idx, column=2, value=s.name)
        rec = attendance_map.get(s.id, {})
        for col_idx, session in enumerate(sessions, 3):
            status = rec.get(session.id, "—")
            cell   = ws2.cell(row=row_idx, column=col_idx, value=status)
            fill   = present_fill if status == "P" else (absent_fill if status == "A" else None)
            styled(cell, fill=fill)
        ws2.row_dimensions[row_idx].height = 18

    # Footer: present count per date
    footer_row = len(students) + 3
    ws2.cell(row=footer_row, column=1, value="Present Count")
    ws2.cell(row=footer_row, column=2, value="")
    styled(ws2.cell(row=footer_row, column=1), fill=summary_fill, font=Font(bold=True))
    for col_idx, session in enumerate(sessions, 3):
        count = sum(
            1 for s in students
            if attendance_map.get(s.id, {}).get(session.id) == "P"
        )
        cell = ws2.cell(row=footer_row, column=col_idx, value=count)
        styled(cell, fill=summary_fill, font=Font(bold=True))

    # Set date column widths
    for col_idx in range(3, 3 + n_date_cols):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 13

    # ── Save and return ───────────────────────────────────────────────────────
    safe_name = subj.name.replace(" ", "_")
    report_path = str(ATTENDANCE_REPORTS_DIR / f"{safe_name}_consolidated.xlsx")
    wb.save(report_path)

    return FileResponse(
        report_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=f"{safe_name}_attendance_consolidated.xlsx"
    )
